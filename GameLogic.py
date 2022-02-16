import pygame
import random
import openpyxl
from PieceFormat import *

# Screen Dimensions.
screen_width = 900
screen_height = 900

# Game Dimensions ( Tetris Board Is 10 Wide By 20 High ).
active_width = 400  # Each Block Is 40 Wide (400 / 10).
active_height = 800  # Each Block Is 40 High (800 / 20).

# Individual Block Size.
square_size = 40

# The Start Location Of The Playing Area Within The Screen Dimensions.
corner_x = (screen_width - active_width) // 2
corner_y = screen_height - active_height - 50


class Piece(object):

    def __init__(self, column, rows, piece):
        self.x = column
        self.y = rows
        self.piece = piece
        self.color = piece_color
        self.rotation = 0


# Function Used To Create / Update The Grid Within The Active Area Of The Game.
def update_grid(previous_position={}):
    # Create A 10x20 Gray Grid.
    grid = [[(191, 191, 191) for _ in range(10)] for _ in range(20)]

    # Parse Array ( Rolls Are Represented By r Columns Represented By c ).
    for r in range(len(grid)):
        for c in range(len(grid[r])):
            # If A Previous Piece Has Achieved A Locked Position, Change Color Of Grid To Said Piece Value.
            if (c, r) in previous_position:
                key = previous_position[(c, r)]
                grid[r][c] = key

    return grid


# Function Used To Draw The Lines Separating The 10x20 Grid Within The Active Area Of The Game.
def draw_grid_line(surface, grid):
    # Vertical Lines.
    for r in range(len(grid)):
        pygame.draw.line(surface, (50, 0, 0), (corner_x, corner_y + r * square_size),
                         (corner_x + active_width, corner_y + r * square_size), 3)
        # Horizontal Lines.
        for c in range(len(grid[r])):
            pygame.draw.line(surface, (50, 0, 0), (corner_x + c * square_size, corner_y),
                             (corner_x + c * square_size, corner_y + active_height), 3)


# Function Used To Maintain The Window Outside The Active Area Of The Game.
def draw_background(surface, grid, score, previous_score, lines):
    # Set The Window Color To Dark Red.
    surface.fill((50, 0, 0))

    # Game Title ( Text Properties ).
    pygame.font.init()
    text = pygame.font.SysFont('helvetica', 60, italic=True)
    info = text.render('Blocks', True, (191, 191, 191))
    # Position Of Tile Within The Surface Background.
    surface.blit(info, (corner_y - 20, active_height - 15))

    # Current User Score.
    text = pygame.font.SysFont('helvetica', 25, italic=True)
    info = text.render('Your Score: ' + str(score), True, (191, 191, 191))
    # Position Of User Score Within The Surface Background.
    surface.blit(info, (corner_y - 20, corner_y))

    # Previous Highest Score.
    info = text.render(previous_score[0] + ', ' + str(previous_score[1]), True, (191, 191, 191))
    # Position Of User Score Within The Surface Background.
    surface.blit(info, (corner_y - 20, corner_y + 40))

    # Current Amount Of Lines Cleared.
    info = text.render('Lines: ' + str(lines), True, (191, 191, 191))
    # Display The Position Of Cleaned Lines By User.
    surface.blit(info, (corner_y - 20, corner_y + 80))

    # Drawing Grid & Border.
    for r in range(len(grid)):
        for c in range(len(grid[r])):
            pygame.draw.rect(surface, grid[r][c], (corner_x + c * square_size, corner_y + r * square_size, square_size,
                                                   square_size), 0)
    draw_grid_line(surface, grid)
    pygame.draw.rect(surface, (50, 0, 0), (corner_x, corner_y, active_width, active_height), 3)


# Function Used To Read The Shape Rotations From The PieceFormat File.
def piece_format_reader(object):
    positions = []
    # Retrieve Shape Based On The Possible Rotations.
    shape = object.piece[object.rotation % len(object.piece)]

    # Parse The Format File And Whenever There Is A "O" Save The Corresponding X & Y Position Of The Piece.
    for r, line in enumerate(shape):
        row = list(line)
        for c, column in enumerate(row):
            if column == 'O':
                positions.append((object.x + c, object.y + r))

    # Allow For The Pieces To Appear Above The Screen From The Start.
    for i, position in enumerate(positions):
        positions[i] = (position[0] - 2, position[1] - 1)
    return positions


# Function Used To Check if The Current Position Of A Piece Does Not Obstruct With Previous Placements.
def check_space(piece, grid):
    # Allow The Adding Of The Position If The Grid At That Location Is Empty, Meaning There Are No Pieces Already There.
    allowed = [[(c, r) for c in range(10) if grid[r][c] == (191, 191, 191)] for r in range(20)]

    # Retrieve The Positions Of The Pieces From The List Above, Instead Of Maintaining The SubLists Created.
    allowed = [c for sub in allowed for c in sub]

    # Read The Current Piece Being Played.
    reader = piece_format_reader(piece)

    # If The Current Piece Position Is Allowed, Meaning It Does Not Obstruct Any Previous Piece Or Is Out Of Bounds.
    for position in reader:
        if position not in allowed:
            if position[1] > - 1:
                return False
    return True


# Function Used To Retrieve A Random Piece From The List Of 7 Possible Options.
def get_piece():
    return Piece(5, 0, random.choice(pieces))


# Function Used To Display The Two Next Shapes In The List To The Player.
def display_next_piece(object, object2, surface):
    # Retrieve Shape Based On The Possible Rotations.
    shape = object.piece[object.rotation % len(object.piece)]
    shape2 = object2.piece[object2.rotation % len(object2.piece)]

    # Text Above The 1st Piece.
    text = pygame.font.SysFont('helvetica', 30, italic=True)
    info = text.render('1st', True, (191, 191, 191))
    surface.blit(info, (active_height - 95, corner_y))
    # Position Of The Next Piece.
    for r, line in enumerate(shape):
        row = list(line)
        for c, column in enumerate(row):
            if column == 'O':
                pygame.draw.rect(surface, (0, 0, 0), (695 + c * square_size, 100 + r * square_size, square_size - 4,
                                                      square_size - 4), 0)

    # Text Above The 2nd Piece.
    text = pygame.font.SysFont('helvetica', 30, italic=True)
    info = text.render('2nd', True, (191, 191, 191))
    surface.blit(info, (active_height - 95, corner_y + 175))
    # Position Of The 2nd Next Piece
    for r, line in enumerate(shape2):
        row = list(line)
        for c, column in enumerate(row):
            if column == 'O':
                pygame.draw.rect(surface, (0, 0, 0), (695 + c * square_size, 275 + r * square_size, square_size - 4,
                                                      square_size - 4), 0)


# Function Used To Check If The User Lost The Game.
def game_loss(positions):
    for position in positions:
        x, y = position
        if y < 1:
            return True
    return False


# Function Used To Ensure That A Line Or Multiple Lines Have Been Filled In Order To Clear it.
def row_completed(grid, previous):
    increment = 0
    # Loop Through Grid Backwards & Delete Cleared Row.
    for r in range(len(grid) - 1, -1, -1):
        row = grid[r]
        if (191, 191, 191) not in row:
            increment += 1
            index = r
            for c in range(len(row)):
                try:
                    del previous[(c, r)]
                except:
                    continue
    # Shift All Rows Down By One.
    if increment > 0:
        for key in sorted(list(previous), key=lambda x: x[1])[:: -1]:
            x, y = key
            if y < index:
                key_2 = (x, y + increment)
                previous[key_2] = previous.pop(key)
    return increment


# Function Used To Display The Top Three User Saved Score From Previous Games Played.
def leader_board():
    # Set The Screen To Be 900 X 600 In Dimension, While The Background Color To Be Dark Red.
    surface = pygame.display.set_mode([900, 600])
    surface.fill((50, 0, 0))
    # Call Function To Save Top Scores And The Index Location From The "Name" List.
    leaderboard, index = top_scores()
    # Writing The Top Three Scores To The Current Surface Along With Other Title Related Options.
    font = pygame.font.SysFont("helvetica", 40, bold=True)
    for top in range(3):
        label = font.render(str(top + 1) + " - " + leaderboard[index[top]][0] + ", Score: "
                            + str(leaderboard[index[top]][1]), True, (191, 191, 191))
        surface.blit(label, (corner_y * 6, corner_y + 125 + (top * 80)))

    font = pygame.font.SysFont("helvetica", 70, bold=True, italic=True)
    label = font.render("Top 3 Scores (Leader Board)", True, (0, 0, 0))
    surface.blit(label, (corner_y + 10, corner_y))

    label = font.render("Press Any Key To Continue!", True, (0, 0, 0))
    surface.blit(label, (corner_y + 10, corner_y + 375))


# Function Used To Determine The Speed & Points of Game Given A User's Line Accomplishments.
def level(lines, score, piece_speed):
    # Level 1
    if lines <= 10:
        score = lines * 10
    # Level 2
    elif lines <= 20:
        back = 100
        temp = lines * 20 - back
        score = temp
        piece_speed = 0.25
    # Level 3
    elif lines <= 30:
        back = 400
        temp = lines * 30 - back
        score = temp
        piece_speed = 0.2
    # Level 4
    elif lines <= 40:
        back = 900
        temp = lines * 40 - back
        score = temp
        piece_speed = 0.15
    # Level 5
    elif lines >= 50:
        back = 1600
        temp = lines * 50 - back
        score = temp
        piece_speed = 0.1
    return score, piece_speed


# Function Used To Create A Box Where Users Will Be Able To Save Their Names To The LeaderBoard.
def user_input():
    background = pygame.display.set_mode([500, 500])

    run2 = True
    active = False
    user_text = ''

    base_font = pygame.font.SysFont("helvetica", 32, italic=True)
    input_rect = pygame.Rect(170, 195, 120, 35)

    color_active = (191, 191, 191)
    color_passive = (0, 0, 0)
    color = color_passive

    while run2:
        for event2 in pygame.event.get():
            if event2.type == pygame.QUIT:
                pygame.quit()
                quit()
            # When The Mouse Button Is Clicked, The Box Will Be Activated Or Deactivated Depending On Conditions.
            if event2.type == pygame.MOUSEBUTTONDOWN:
                # If The Mouse Click Collides With The Box On The Screen, The Box Will be Activated.
                if input_rect.collidepoint(event2.pos):
                    active = True
                else:
                    # If User Clicks Out Of The Box & The Input Its Empty, The Box Will Be Deactivated.
                    if user_text == '':
                        active = False
                    # Else, When Users Click Out Of The Box & There Is Some Input, The Loop Will Stop Running.
                    else:
                        run2 = False
            # When A Keyboard Key Is Pressed, User Keys Will Be Recorded.
            if event2.type == pygame.KEYDOWN:
                if active == True:
                    # Allows Users To Use Backspace.
                    if event2.key == pygame.K_BACKSPACE:
                        user_text = user_text[:-1]
                    # Appends User Keyboard Unicode Values To The String "User_Text" Created Above.
                    else:
                        user_text += event2.unicode

        background.fill((50, 0, 0))

        # Writing Information To The Screen, For Example, The "Thank You For Playing" Message.
        font = pygame.font.SysFont("helvetica", 50, bold=True, italic=True)
        label = font.render("Thank You For Playing!", True, (0, 0, 0))
        background.blit(label, (corner_y - 30, corner_y))

        font = pygame.font.SysFont("helvetica", 20, bold=True, italic=True)
        label = font.render("Click On The Square & Write Your Name", True, (191, 191, 191))
        background.blit(label, (corner_y + 45, corner_y + 85))
        label = font.render("Click Outside Of The Square To Save Score", True, (191, 191, 191))
        background.blit(label, (corner_y + 30, corner_y + 210))

        font = pygame.font.SysFont("helvetica", 25, bold=True, italic=True)
        label = font.render("Additional Thanks To,", True, (0, 0, 0))
        background.blit(label, (corner_y + 90, corner_y + 270))
        label = font.render("FreeCodeCamp & Tech With Tim", True, (0, 0, 0))
        background.blit(label, (corner_y + 40, corner_y + 310))
        label = font.render("(Pygame Course)", True, (0, 0, 0))
        background.blit(label, (corner_y + 110, corner_y + 350))

        # If Active = True, Change Box To Gray.
        if active:
            color = color_active
        # Else, Change Box To Black.
        else:
            color = color_passive

        # Drawing The Rectangle To The Screen & Allowing The Width To Be Changed Based On The Length of The String.
        pygame.draw.rect(background, color, input_rect)
        text_surface = base_font.render(user_text, True, (0, 0, 0))
        background.blit(text_surface, (input_rect.x + 5, input_rect.y))
        input_rect.w = max(160, text_surface.get_width() + 10)

        pygame.display.update()
    return user_text


# Function Used To Read The LeaderBoard File.
def read():
    file = openpyxl.load_workbook('LeaderBoard.xlsx')
    page = file['Sheet']

    row = page.max_row
    column = page.max_column

    leaderboard = []

    for i in range(1, row + 1):
        line = []
        for j in range(1, column + 1):
            line.append(page.cell(i, j).value)
        leaderboard.append(line)
    return leaderboard


# Function Used To Return The Leaderboard (Name, Score) & The Index Of The Top Numbers From That List.
def top_scores():
    leaderboard = read()
    # Pop The First Sublist From The Main List, (Name, Score)
    leaderboard.pop(0)
    top = []
    index = []

    for i in range(len(leaderboard)):
        # Append All The Scores Into A List.
        top.append(leaderboard[i][1])
    for i in range(len(top)):
        # Get The Highest Value In The List.
        greater = max(top)
        # Get The Index Of The Highest Value From list.
        greater_index = top.index(greater)
        # Set The Highest Value In The List To Zero.
        top[greater_index] = 0
        # Append The Index To Another List.
        index.append(greater_index)
    return leaderboard, index


# Function Used To Write The UserName & Score To An Excel File.
def write(score, text):
    leaderboard = read()
    leaderboard.append([text, score])
    file2 = openpyxl.Workbook()
    page2 = file2.active
    # Loop Used To Fill The Cells In A Way Where Users & Scores Will Have Distinct Columns But Sharing Same Rows.
    for row in range(len(leaderboard)):
        for column in range(len(leaderboard[0])):
            page2.cell(row + 1, column + 1).value = leaderboard[row][column]
    file2.save('LeaderBoard.xlsx')

